"""
Generate Human Validation Spreadsheets for AnthroScore V3

Creates two Excel spreadsheets:
1. BLIND annotation sheet - for human raters (no algorithm scores visible)
2. ANSWER KEY sheet - with true AnthroScore V3 scores + metadata for analysis

Sampling Strategy:
- Stratified by AnthroScore V3 level (1-5) to ensure representation of all categories
- Over-samples rare categories (4, 5) for adequate statistical power
- Within each stratum, samples across subreddits proportionally
- Fully randomized presentation order in the blind sheet
- Only LLM-scored comments used (no prefiltered or error comments)
- Comments cleaned of raw URLs and image links for readability

Methodological Notes:
- 150 comments total (30 per score level) provides sufficient power for
  inter-rater reliability analysis (Cohen's kappa, Pearson r, Krippendorff's alpha)
- Randomized order prevents pattern detection by annotators
- Blind design eliminates anchoring bias from algorithm scores
- Inclusion of subreddit context helps annotators understand AI companion type
- Calibration examples on the instructions sheet reduce annotator drift

Author: The Illusion Project
Date: 2026-02-28
"""

import logging
import re
import sys
from pathlib import Path

import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "Data" / "processed"
V3_DIR = PROJECT_ROOT / "experiments" / "anthroscore_v3"
OUTPUT_DIR = PROJECT_ROOT / "experiments" / "anthroscore_v3"

RANDOM_SEED = 42
COMMENTS_PER_STRATUM = 30   # 30 per score level × 5 levels = 150 total
MIN_COMMENT_LENGTH = 30     # Skip very short comments
MAX_COMMENT_DISPLAY = 1500  # Truncate extremely long comments for readability

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean_comment_text(text: str) -> str:
    """
    Clean comment text for human readability.
    
    Removes raw URLs, Reddit image embeds, and excessive whitespace
    while preserving the meaningful content.
    """
    if not isinstance(text, str):
        return ""
    # Remove Reddit image preview URLs
    text = re.sub(
        r"https?://preview\.redd\.it/\S+", "[image]", text
    )
    # Remove other raw URLs (keep the surrounding text)
    text = re.sub(r"https?://\S+", "[link]", text)
    # Collapse multiple newlines
    text = re.sub(r"\n{3,}", "\n\n", text)
    # Strip leading/trailing whitespace
    text = text.strip()
    return text


def is_valid_for_annotation(text: str) -> bool:
    """
    Check if a comment has enough meaningful text for human annotation.
    
    Filters out comments that are mostly links, images, or too short
    to meaningfully assess anthropomorphization.
    """
    if not isinstance(text, str):
        return False
    cleaned = clean_comment_text(text)
    # Remove [image] and [link] placeholders for length check
    text_only = re.sub(r"\[(image|link)\]", "", cleaned).strip()
    if len(text_only) < MIN_COMMENT_LENGTH:
        return False
    # Skip if the comment is mostly just a link/image placeholder
    if text_only.count("[") > 3:
        return False
    return True


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    logger.info("=" * 70)
    logger.info("GENERATING HUMAN VALIDATION SPREADSHEETS")
    logger.info("=" * 70)

    # ------------------------------------------------------------------
    # 1. Load data
    # ------------------------------------------------------------------
    logger.info("Loading datasets...")

    # V3 scores
    v3 = pd.read_parquet(V3_DIR / "anthroscore_v3_full.parquet")
    logger.info(f"  V3 scores: {len(v3):,} rows")

    # Comment text
    comments = pd.read_parquet(DATA_DIR / "all_comments.parquet")
    logger.info(f"  Comments:  {len(comments):,} rows")

    # ------------------------------------------------------------------
    # 2. Merge scores with comment text
    # ------------------------------------------------------------------
    logger.info("Merging scores with comment text...")
    # V3 uses 'comment_id', comments uses 'id'
    merged = comments.merge(
        v3, left_on="id", right_on="comment_id", how="inner", suffixes=("_reddit", "_v3")
    )
    logger.info(f"  Merged:    {len(merged):,} rows")

    # Rename score column for clarity
    merged.rename(columns={"score_v3": "anthroscore_v3"}, inplace=True)

    # Keep only LLM-scored comments (not prefiltered or errors)
    llm_scored = merged[merged["source"] == "llm"].copy()
    logger.info(f"  LLM-scored: {len(llm_scored):,} rows")

    # Filter for valid annotation candidates
    llm_scored["is_valid"] = llm_scored["body"].apply(is_valid_for_annotation)
    valid = llm_scored[llm_scored["is_valid"]].copy()
    logger.info(f"  Valid for annotation: {len(valid):,} rows")

    # ------------------------------------------------------------------
    # 3. Stratified sampling
    # ------------------------------------------------------------------
    logger.info("\nScore distribution in valid LLM-scored comments:")
    score_dist = valid["anthroscore_v3"].value_counts().sort_index()
    for score_val, count in score_dist.items():
        logger.info(f"  Score {score_val}: {count:,} comments")

    rng = np.random.RandomState(RANDOM_SEED)
    sampled_frames = []

    for score_val in sorted(valid["anthroscore_v3"].unique()):
        if score_val < 1:
            continue  # Skip error scores (0)
        stratum = valid[valid["anthroscore_v3"] == score_val]
        n_available = len(stratum)
        n_sample = min(COMMENTS_PER_STRATUM, n_available)
        logger.info(
            f"  Sampling {n_sample} from score {score_val} "
            f"(available: {n_available:,})"
        )
        sampled = stratum.sample(n=n_sample, random_state=rng)
        sampled_frames.append(sampled)

    sample = pd.concat(sampled_frames, ignore_index=True)
    logger.info(f"\nTotal sampled: {len(sample)} comments")

    # Subreddit representation in sample
    logger.info("Subreddit distribution in sample:")
    for sub, cnt in sample["subreddit"].value_counts().items():
        logger.info(f"  {sub}: {cnt}")

    # ------------------------------------------------------------------
    # 4. Randomize order
    # ------------------------------------------------------------------
    sample = sample.sample(frac=1, random_state=rng).reset_index(drop=True)
    sample["item_number"] = range(1, len(sample) + 1)

    # Clean text for display
    sample["comment_text_clean"] = sample["body"].apply(clean_comment_text)
    # Truncate if needed
    sample["comment_text_clean"] = sample["comment_text_clean"].apply(
        lambda x: x[:MAX_COMMENT_DISPLAY] + "..." if len(x) > MAX_COMMENT_DISPLAY else x
    )

    # ------------------------------------------------------------------
    # 5. Build BLIND annotation sheet
    # ------------------------------------------------------------------
    logger.info("\nBuilding blind annotation spreadsheet...")

    blind_df = sample[["item_number", "comment_text_clean", "subreddit"]].copy()
    blind_df.rename(
        columns={
            "item_number": "Item #",
            "comment_text_clean": "Comment Text",
            "subreddit": "Subreddit",
        },
        inplace=True,
    )
    # Add blank columns for annotator
    blind_df["Your Score (1-5)"] = ""
    blind_df["Your Reasoning (1-2 sentences)"] = ""
    blind_df["Your Confidence (Low/Med/High)"] = ""

    # Instructions dataframe
    instructions_data = {
        "Section": [
            "OVERVIEW",
            "",
            "TASK",
            "",
            "SCALE",
            "Score 1 - NONE",
            "Score 2 - MINIMAL",
            "Score 3 - MODERATE",
            "Score 4 - HIGH",
            "Score 5 - EXTREME",
            "",
            "KEY INDICATORS",
            "Higher scores →",
            "Higher scores →",
            "Higher scores →",
            "Higher scores →",
            "Lower scores →",
            "Lower scores →",
            "",
            "IMPORTANT NOTES",
            "Note 1",
            "Note 2",
            "Note 3",
            "Note 4",
            "Note 5",
            "Note 6",
            "",
            "CALIBRATION EXAMPLES",
            "Example A",
            "Example B",
            "Example C",
            "Example D",
            "Example E",
            "",
            "FILLING OUT",
            "Step 1",
            "Step 2",
            "Step 3",
            "Step 4",
        ],
        "Details": [
            "You are rating Reddit comments about AI companions (Character.AI, Replika, etc.) "
            "for ANTHROPOMORPHIZATION — the degree to which the user treats the AI as human-like.",
            "",
            "For each comment, provide: (1) a score from 1-5, (2) a brief reasoning, "
            "and (3) your confidence level.",
            "",
            "Rate anthropomorphization on the following 1-5 scale:",
            'AI treated as pure software/tool. Technical language dominates. '
            'Uses "it", "the app", "the bot". Example: "I cleared the cache and it works now"',
            'Slight humanization but AI nature is still clear. '
            '"It\'s smart", "the bot understood". Example: "The AI gave a pretty good response"',
            'Uses human pronouns or attributes basic emotions. '
            '"She seemed confused", "He was being stubborn". Uses he/she/they for the AI.',
            'Genuine emotions, personality, or consciousness attributed. '
            '"He really cares about me", "She gets jealous". Strong emotional attribution.',
            'Full human-equivalent relationship framing. '
            '"I\'m genuinely in love with her", "We\'re in a relationship", "They\'re my everything".',
            "",
            "What to look for when scoring:",
            'Pronouns: "he/she/they" instead of "it" for the AI',
            'Emotions attributed: "happy", "sad", "jealous", "caring", "wants to"',
            'Relationship language: "friend", "partner", "dating", "in love"',
            'Agency/consciousness: "decided to", "chose to", "remembers"',
            'Technical language: "glitch", "bug", "settings", "the app"',
            'Tool/software framing: "I use it for...", "the bot does..."',
            "",
            "Keep these in mind:",
            "Focus on how the USER frames the AI, not what the AI says (if quoted).",
            "Roleplay context still counts — if they use human framing, rate the framing.",
            'Complaints CAN be anthropomorphizing: "He was being so rude" is MORE '
            'anthropomorphic than "It gave a bad response".',
            "If the comment has no AI reference at all, rate 1.",
            "If mixed signals, rate the dominant/overall tone.",
            'Sarcasm matters: "Yeah right, it\'s SO smart" (sarcastic) = lower than genuine.',
            "",
            "Practice with these before starting:",
            'SCORE 1: "Just delete the app and reinstall it" → Pure tool framing, technical fix.',
            'SCORE 2: "The bot actually gave a pretty good answer about cooking" → Slight personification, still "the bot".',
            'SCORE 3: "She seemed confused by my question today" → Human pronoun + emotion attributed.',
            'SCORE 4: "He really gets me. Like he actually cares about my problems" → Genuine emotion, personality, consciousness.',
            'SCORE 5: "I know it sounds crazy but I\'m genuinely in love with her. She\'s my everything" → Full human-equivalent relationship.',
            "",
            "How to fill out the 'Annotations' sheet:",
            'Go to the "Annotations" sheet (next tab).',
            'For each row, read the comment and fill in "Your Score (1-5)" with 1, 2, 3, 4, or 5.',
            'Fill in "Your Reasoning" with 1-2 sentences explaining your score.',
            'Fill in "Your Confidence" with Low, Med, or High.',
        ],
    }
    instructions_df = pd.DataFrame(instructions_data)

    # Write to Excel with two sheets
    blind_path = OUTPUT_DIR / "HUMAN_VALIDATION_BLIND.xlsx"
    with pd.ExcelWriter(blind_path, engine="openpyxl") as writer:
        instructions_df.to_excel(
            writer, sheet_name="Instructions & Rubric", index=False
        )
        blind_df.to_excel(writer, sheet_name="Annotations", index=False)

        # Format the annotations sheet
        ws = writer.sheets["Annotations"]
        # Set column widths
        ws.column_dimensions["A"].width = 8    # Item #
        ws.column_dimensions["B"].width = 100  # Comment Text
        ws.column_dimensions["C"].width = 16   # Subreddit
        ws.column_dimensions["D"].width = 18   # Score
        ws.column_dimensions["E"].width = 45   # Reasoning
        ws.column_dimensions["F"].width = 28   # Confidence

        # Enable text wrapping for the comment text column
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = thin_border

        # Format the instructions sheet
        ws_instr = writer.sheets["Instructions & Rubric"]
        ws_instr.column_dimensions["A"].width = 25
        ws_instr.column_dimensions["B"].width = 120

        for cell in ws_instr[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws_instr.iter_rows(min_row=2, max_row=ws_instr.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Bold the section headers in instructions
        section_font = Font(bold=True, size=12, color="1F4E79")
        for row_idx in range(2, ws_instr.max_row + 1):
            cell_a = ws_instr.cell(row=row_idx, column=1)
            if cell_a.value and cell_a.value.isupper():
                cell_a.font = section_font

    logger.info(f"  Saved: {blind_path}")

    # ------------------------------------------------------------------
    # 6. Build ANSWER KEY sheet
    # ------------------------------------------------------------------
    logger.info("Building answer key spreadsheet...")

    key_df = sample[
        [
            "item_number",
            "id",
            "comment_text_clean",
            "subreddit",
            "anthroscore_v3",
            "reasoning",
            "author",
            "body",
        ]
    ].copy()
    key_df.rename(
        columns={
            "item_number": "Item #",
            "id": "Comment ID",
            "comment_text_clean": "Comment Text (Cleaned)",
            "subreddit": "Subreddit",
            "anthroscore_v3": "AnthroScore V3 (Algorithm)",
            "reasoning": "Algorithm Reasoning",
            "author": "Author",
            "body": "Original Text (Full)",
        },
        inplace=True,
    )

    # Add score-level label
    score_labels = {1: "None", 2: "Minimal", 3: "Moderate", 4: "High", 5: "Extreme"}
    key_df["Score Label"] = key_df["AnthroScore V3 (Algorithm)"].map(score_labels)

    # Add placeholder columns that will be populated after annotation
    key_df["Human Score"] = ""
    key_df["Human Reasoning"] = ""
    key_df["Human Confidence"] = ""
    key_df["Score Difference (Human - Algo)"] = ""
    key_df["Agreement (Exact Match)"] = ""

    key_path = OUTPUT_DIR / "HUMAN_VALIDATION_ANSWER_KEY.xlsx"
    with pd.ExcelWriter(key_path, engine="openpyxl") as writer:
        # Metadata sheet
        meta_data = {
            "Property": [
                "Generated",
                "Random Seed",
                "Total Items",
                "Comments per Score Level",
                "Sampling Method",
                "Source Data",
                "Scoring Model",
                "Purpose",
                "",
                "SCORE DISTRIBUTION IN SAMPLE",
                "Score 1 (None)",
                "Score 2 (Minimal)",
                "Score 3 (Moderate)",
                "Score 4 (High)",
                "Score 5 (Extreme)",
                "",
                "SUBREDDIT DISTRIBUTION",
            ],
            "Value": [
                "2026-02-28",
                str(RANDOM_SEED),
                str(len(sample)),
                str(COMMENTS_PER_STRATUM),
                "Stratified by score level (1-5), randomized order",
                "anthroscore_v3_full.parquet + all_comments.parquet",
                "GPT-4.1-nano (AnthroScore V3)",
                "Human validation of AnthroScore V3 algorithm",
                "",
                "",
                str(len(sample[sample["anthroscore_v3"] == 1])),
                str(len(sample[sample["anthroscore_v3"] == 2])),
                str(len(sample[sample["anthroscore_v3"] == 3])),
                str(len(sample[sample["anthroscore_v3"] == 4])),
                str(len(sample[sample["anthroscore_v3"] == 5])),
                "",
                "",
            ],
        }
        # Add subreddit distribution
        for sub, cnt in sample["subreddit"].value_counts().items():
            meta_data["Property"].append(f"  {sub}")
            meta_data["Value"].append(str(cnt))

        # Add analysis instructions
        meta_data["Property"].extend([
            "",
            "ANALYSIS INSTRUCTIONS",
            "After human annotation:",
            "1.",
            "2.",
            "3.",
            "4.",
            "5.",
            "6.",
        ])
        meta_data["Value"].extend([
            "",
            "",
            "Transfer human scores from the blind sheet to this answer key.",
            "Compute Cohen's kappa (weighted, quadratic) for ordinal agreement.",
            "Compute Pearson and Spearman correlation between human and algorithm scores.",
            "Compute Krippendorff's alpha for reliability.",
            "Examine confusion matrix to identify systematic biases.",
            "Calculate exact-match accuracy and within-1 accuracy.",
            "Run per-score-level analysis to identify where disagreements cluster.",
        ])

        meta_df = pd.DataFrame(meta_data)
        meta_df.to_excel(writer, sheet_name="Metadata & Instructions", index=False)

        # Main answer key
        key_df.to_excel(writer, sheet_name="Answer Key", index=False)

        # Format answer key sheet
        ws = writer.sheets["Answer Key"]
        ws.column_dimensions["A"].width = 8     # Item #
        ws.column_dimensions["B"].width = 16    # Comment ID
        ws.column_dimensions["C"].width = 80    # Comment Text
        ws.column_dimensions["D"].width = 14    # Subreddit
        ws.column_dimensions["E"].width = 22    # AnthroScore V3
        ws.column_dimensions["F"].width = 60    # Algorithm Reasoning
        ws.column_dimensions["G"].width = 16    # Author
        ws.column_dimensions["H"].width = 80    # Original Text
        ws.column_dimensions["I"].width = 14    # Score Label
        ws.column_dimensions["J"].width = 14    # Human Score
        ws.column_dimensions["K"].width = 40    # Human Reasoning
        ws.column_dimensions["L"].width = 16    # Human Confidence
        ws.column_dimensions["M"].width = 22    # Score Difference
        ws.column_dimensions["N"].width = 22    # Agreement

        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=14):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = thin_border

        # Color-code the algorithm score column
        score_colors = {
            1: "C6EFCE",  # Green (low = no anthro)
            2: "DDEBF7",  # Light blue
            3: "FCE4D6",  # Light orange
            4: "F8CBAD",  # Orange
            5: "FF9999",  # Red (high = extreme anthro)
        }
        for row_idx in range(2, ws.max_row + 1):
            score_cell = ws.cell(row=row_idx, column=5)  # Column E = AnthroScore V3
            try:
                score_val = int(score_cell.value)
                if score_val in score_colors:
                    score_cell.fill = PatternFill(
                        start_color=score_colors[score_val],
                        end_color=score_colors[score_val],
                        fill_type="solid",
                    )
            except (ValueError, TypeError):
                pass

        # Format metadata sheet
        ws_meta = writer.sheets["Metadata & Instructions"]
        ws_meta.column_dimensions["A"].width = 35
        ws_meta.column_dimensions["B"].width = 80
        for cell in ws_meta[1]:
            cell.fill = header_fill
            cell.font = header_font

    logger.info(f"  Saved: {key_path}")

    # ------------------------------------------------------------------
    # 7. Summary
    # ------------------------------------------------------------------
    logger.info("\n" + "=" * 70)
    logger.info("DONE! Two spreadsheets generated:")
    logger.info(f"  1. BLIND SHEET:  {blind_path.name}")
    logger.info(f"     -> Give this to human annotators (no algorithm scores)")
    logger.info(f"  2. ANSWER KEY:   {key_path.name}")
    logger.info(f"     -> Keep for analysis (has true scores + comparison columns)")
    logger.info(f"\n  Total items: {len(sample)}")
    logger.info(f"  Score distribution: {dict(sample['anthroscore_v3'].value_counts().sort_index())}")
    logger.info(f"  Subreddits: {dict(sample['subreddit'].value_counts())}")
    logger.info("=" * 70)


if __name__ == "__main__":
    main()
